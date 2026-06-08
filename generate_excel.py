import sys, json, re, unicodedata
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

STOP_WORDS = {'quel','quelle','quels','quelles','est','sont','avez','vous','votre','vos','etes','avons','ont','avoir','quoi','comment','combien','pourquoi','si','oui','non','les','des','une','un','la','le','au','aux','du','de','que','qui','quand','par','pour','avec','sans','sur','sous','dans','entre','vers','chez','en','et','mais','donc','car','ni','or','pas','ne','plus','moins','tres','bien','tout','tous','toute','toutes','autre','autres','meme','plusieurs','quelques','ete','fait','fois','avait','etait','sera','serait','avoir','etre','faire','dire','aller'}

def make_var_name(label):
    label = unicodedata.normalize('NFD', label)
    label = ''.join(c for c in label if unicodedata.category(c) != 'Mn')
    label = label.lower()
    words = re.sub(r'[^a-z0-9\s]', ' ', label).split()
    words = [w for w in words if len(w) > 2 and w not in STOP_WORDS]
    var = '_'.join(words[:3])
    return var[:30].rstrip('_') if var else 'variable'

def generate(form_json, output_path):
    form = json.loads(form_json)
    questions = form.get('questions', [])

    # Garantir l'unicité des noms
    used = {}
    col_meta = []
    groups = {}
    group_order = []
    for q in questions:
        g = q.get('group') or 'General'
        if g not in groups:
            groups[g] = []
            group_order.append(g)
        groups[g].append(q)
        base = make_var_name(q.get('label', q.get('id', 'var')))
        if base not in used:
            used[base] = 1
            varname = base
        else:
            used[base] += 1
            varname = base + '_' + str(used[base])
        col_meta.append({'q': q, 'group': g, 'varname': varname})

    wb = openpyxl.Workbook()

    # Styles
    FILL_SEC = PatternFill("solid", fgColor="1F4E79")
    FILL_VAR = PatternFill("solid", fgColor="2E75B6")
    FILL_ODD = PatternFill("solid", fgColor="EBF3FB")
    FILL_EVN = PatternFill("solid", fgColor="FFFFFF")
    FNT_WH  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    FNT_VAR = Font(name='Arial', bold=True, color='FFFFFF', size=9)
    FNT_DAT = Font(name='Arial', size=9)
    BRD = Border(
        left=Side(style='thin', color='BDD7EE'),
        right=Side(style='thin', color='BDD7EE'),
        top=Side(style='thin', color='BDD7EE'),
        bottom=Side(style='thin', color='BDD7EE')
    )
    CTR = Alignment(horizontal='center', vertical='center')
    LFT = Alignment(horizontal='left', vertical='center')

    # ====== FEUILLE 1: SAISIE ======
    ws1 = wb.active
    ws1.title = "Saisie"

    # Ligne 1: sections fusionnées
    col = 1
    for g in group_order:
        n = len(groups[g])
        c = ws1.cell(row=1, column=col, value=g)
        c.font = FNT_WH; c.fill = FILL_SEC; c.alignment = CTR; c.border = BRD
        if n > 1:
            ws1.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+n-1)
        col += n
    ws1.row_dimensions[1].height = 22

    # Ligne 2: noms variables
    for i, meta in enumerate(col_meta):
        c = ws1.cell(row=2, column=i+1, value=meta['varname'])
        c.font = FNT_VAR; c.fill = FILL_VAR; c.alignment = CTR; c.border = BRD
    ws1.row_dimensions[2].height = 20
    ws1.freeze_panes = 'A3'

    # 100 lignes de saisie
    N = 100
    for row in range(3, 3+N):
        fill = FILL_ODD if row % 2 == 1 else FILL_EVN
        for i in range(len(col_meta)):
            c = ws1.cell(row=row, column=i+1, value='')
            c.font = FNT_DAT; c.fill = fill; c.border = BRD; c.alignment = LFT

    # Validations et formats
    for i, meta in enumerate(col_meta):
        q = meta['q']
        t = q.get('selectedType') or q.get('type', 'text')
        cl = get_column_letter(i+1)
        rng = f"{cl}3:{cl}{2+N}"

        if t in ('select_one', 'select_multiple'):
            choices = q.get('choices', [])
            if choices:
                labels = [str(c) if isinstance(c, str) else c.get('label', str(c)) for c in choices]
                formula = '"' + ','.join(labels[:20]) + '"'
                if len(formula) <= 257:
                    dv = DataValidation(type='list', formula1=formula, showDropDown=False,
                                       allow_blank=not q.get('required', True))
                    dv.error = 'Choisissez une valeur dans la liste'
                    dv.errorTitle = 'Valeur invalide'
                    dv.prompt = 'Selectionnez une option'
                    dv.promptTitle = (q.get('label') or '')[:32]
                    dv.sqref = rng
                    ws1.add_data_validation(dv)

        elif t == 'integer':
            nm, nx = q.get('numMin', ''), q.get('numMax', '')
            nm_ok = nm not in ('', None)
            nx_ok = nx not in ('', None)
            if nm_ok and nx_ok:
                dv = DataValidation(type='whole', operator='between',
                                   formula1=str(nm), formula2=str(nx), allow_blank=True)
                dv.error = f'Entrez un entier entre {nm} et {nx}'
            elif nm_ok:
                dv = DataValidation(type='whole', operator='greaterThanOrEqual',
                                   formula1=str(nm), allow_blank=True)
                dv.error = f'Entrez un entier >= {nm}'
            elif nx_ok:
                dv = DataValidation(type='whole', operator='lessThanOrEqual',
                                   formula1=str(nx), allow_blank=True)
                dv.error = f'Entrez un entier <= {nx}'
            else:
                dv = DataValidation(type='whole', operator='between',
                                   formula1='-999999', formula2='999999', allow_blank=True)
                dv.error = 'Entrez un nombre entier'
            dv.errorTitle = 'Valeur invalide'
            dv.sqref = rng
            ws1.add_data_validation(dv)
            for row in range(3, 3+N):
                ws1.cell(row=row, column=i+1).number_format = '0'

        elif t == 'decimal':
            nm, nx = q.get('numMin', ''), q.get('numMax', '')
            nm_ok = nm not in ('', None)
            nx_ok = nx not in ('', None)
            if nm_ok and nx_ok:
                dv = DataValidation(type='decimal', operator='between',
                                   formula1=str(nm), formula2=str(nx), allow_blank=True)
            else:
                dv = DataValidation(type='decimal', operator='between',
                                   formula1='-999999', formula2='999999', allow_blank=True)
            dv.error = 'Entrez un nombre decimal valide'
            dv.errorTitle = 'Valeur invalide'
            dv.sqref = rng
            ws1.add_data_validation(dv)
            after = q.get('numDigitsAfter', '')
            fmt = '0.' + '0' * (int(after) if str(after).isdigit() else 2)
            for row in range(3, 3+N):
                ws1.cell(row=row, column=i+1).number_format = fmt

        elif t == 'date':
            dv = DataValidation(type='date', allow_blank=True)
            dv.prompt = 'Format: JJ/MM/AAAA'
            dv.promptTitle = 'Date'
            dv.sqref = rng
            ws1.add_data_validation(dv)
            for row in range(3, 3+N):
                ws1.cell(row=row, column=i+1).number_format = 'DD/MM/YYYY'

        elif t == 'time':
            for row in range(3, 3+N):
                ws1.cell(row=row, column=i+1).number_format = 'HH:MM'

        elif t == 'datetime':
            for row in range(3, 3+N):
                ws1.cell(row=row, column=i+1).number_format = 'DD/MM/YYYY HH:MM'

        # Largeur colonne
        ws1.column_dimensions[cl].width = max(15, min(28, len(meta['varname'])+4))

    # ====== FEUILLE 2: DICTIONNAIRE ======
    ws2 = wb.create_sheet("Dictionnaire")
    hdrs = ['Variable', 'Libelle complet de la question', 'Type', 'Section', 'Obligatoire', 'Modalites']
    for j, h in enumerate(hdrs):
        c = ws2.cell(row=1, column=j+1, value=h)
        c.font = FNT_WH; c.fill = FILL_SEC; c.alignment = CTR; c.border = BRD
    ws2.row_dimensions[1].height = 20
    ws2.freeze_panes = 'A2'

    type_labels = {
        'select_one':'Choix unique','select_multiple':'Choix multiple',
        'integer':'Nombre entier','decimal':'Nombre decimal',
        'date':'Date','time':'Heure','datetime':'Date et heure',
        'text':'Texte libre','calculate':'Calcul automatique',
        'geopoint':'GPS','image':'Photo','audio':'Audio',
        'video':'Video','file':'Fichier','barcode':'Code-barres',
        'acknowledge':'Confirmation','rank':'Classement','range':'Echelle','note':'Note'
    }

    for r, meta in enumerate(col_meta):
        q = meta['q']
        t = q.get('selectedType') or q.get('type', 'text')
        choices = q.get('choices', [])
        choice_str = ' / '.join([str(c) if isinstance(c, str) else c.get('label', '') for c in choices])
        fill = FILL_ODD if r % 2 == 0 else FILL_EVN
        row_data = [
            meta['varname'],
            q.get('label', ''),
            type_labels.get(t, t),
            meta['group'],
            'Oui' if q.get('required', True) else 'Non',
            choice_str
        ]
        for j, val in enumerate(row_data):
            c = ws2.cell(row=r+2, column=j+1, value=val)
            c.font = FNT_DAT; c.fill = fill; c.border = BRD
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=(j in (1, 5)))

    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 55
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 25
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 45

    wb.save(output_path)

if __name__ == '__main__':
    generate(sys.argv[1], sys.argv[2])
    print('ok')
